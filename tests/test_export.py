import unittest

import filecmp
from openpyxl import load_workbook
from pandas.util.testing import assert_frame_equal

from pymongo_schema.export import *
from tests.tools import TestRemovingOutputOnSuccess

TEST_DIR = os.path.dirname(__file__)


class TestExportUnit(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.simple_schema = {'count': 25359,
                             'object': {'field': {'types_count': {'string': 25359},
                                                  'count': 25359, 'type': 'string',
                                                  'prop_in_object': 1.0}}}
        cls.long_schema = {
            'count': 25359,
            'object': {'field': {'types_count': {'string': 25359}, 'count': 25359, 'type': 'string',
                                 'prop_in_object': 1.0},
                       'field2': {'types_count': {'string': 25359}, 'count': 25359,
                                  'type': 'string',
                                  'prop_in_object': 1.0},
                       'field3': {'types_count': {'ARRAY': 25359}, 'count': 25359, 'type': 'ARRAY',
                                  'prop_in_object': 1.0, 'array_types_count': {'OBJECT': 25359},
                                  'array_type': 'OBJECT',
                                  'object': {'subfield1': {'types_count': {'string': 25359},
                                                           'count': 25359, 'type': 'string',
                                                           'prop_in_object': 1.0},
                                             'subfield2': {'types_count': {'object': 25359},
                                                           'count': 25359, 'type': 'ARRAY',
                                                           'prop_in_object': 1.0,
                                                           'array_type': 'string',
                                                           'array_types_count': {'string': 93463,
                                                                                 'null': 738}}}}}}
        cls.simple_full_schema = {'db': {'coll': cls.simple_schema}}
        cls.long_full_schema = {'db1': {'coll': cls.long_schema},
                                'db2': {'coll1': cls.simple_schema, 'coll2': cls.simple_schema}}

    def test00_field_compact_name(self):
        self.assertEqual(field_compact_name('baz', None, 'foo.bar:'), ' .  : baz')
        self.assertEqual(field_compact_name('baz', None, 'foo.foo.bar:'), ' .  .  : baz')

    def test01_field_depth(self):
        self.assertEqual(field_depth(None, None, ''), 0)
        self.assertEqual(field_depth(None, None, 'foo'), 0)
        self.assertEqual(field_depth(None, None, 'foo.bar'), 1)
        self.assertEqual(field_depth(None, None, 'foo.bar:'), 2)
        self.assertEqual(field_depth(None, None, 'foo.foo.bar:'), 3)

    def test02_field_type(self):
        self.assertEqual(field_type(None, {'type': 'string'}, None), 'string')
        self.assertEqual(field_type(None, {'type': 'ARRAY', 'array_type': 'string'}, None),
                         'ARRAY(string)')

    def test03_format_types_count(self):
        res = format_types_count({'integer': 10, 'boolean': 5, 'null': 3, })
        self.assertEqual(res, 'integer : 10, boolean : 5, null : 3')
        res = format_types_count({'ARRAY': 10, 'null': 3, }, {'float': 4})
        self.assertEqual(res, 'ARRAY(float : 4) : 10, null : 3')

    def test04_remove_counts_from_schema_simple(self):
        schema = self.simple_schema
        exp = {'object': {'field': {'type': 'string'}}}
        self.assertEqual(remove_counts_from_schema(schema), exp)

    def test05_remove_counts_from_schema_empty(self):
        self.assertEqual(remove_counts_from_schema({}), {})

    def test06_remove_counts_from_schema_long(self):
        schema = self.long_schema
        exp = {'object': {'field': {'type': 'string'},
                          'field2': {'type': 'string'},
                          'field3': {'type': 'ARRAY', 'array_type': 'OBJECT',
                                     'object': {'subfield1': {'type': 'string'},
                                                'subfield2': {'type': 'ARRAY',
                                                              'array_type': 'string'}}}}}
        self.assertEqual(remove_counts_from_schema(schema), exp)

    def test07_field_schema_to_columns_simple(self):
        res = field_schema_to_columns('field', self.simple_schema['object']['field'], '',
                                      ['Field_compact_name', 'Field_name',
                                       'Count', 'Percentage', 'Types_count'])
        self.assertEqual(res, ('field', 'field', 25359, 100.0, 'string : 25359'))

    def test08_field_schema_to_columns_long(self):
        res = field_schema_to_columns('field3', self.long_schema['object']['field3'], 'foo.bar:',
                                      ['Field_full_name', 'Field_compact_name', 'Field_name',
                                       'Depth', 'Type', 'Count', 'Percentage', 'Types_count'])
        exp = ('foo.bar:field3', ' .  : field3', 'field3', 2, 'ARRAY(OBJECT)', 25359, 100.0,
               'ARRAY(OBJECT : 25359) : 25359')
        self.assertEqual(res, exp)

    def test09_object_schema_to_line_tuples_simple(self):
        res = object_schema_to_line_tuples(
            self.simple_schema['object'], ['Field_compact_name', 'Types_count'], '')
        self.assertEqual(res, [('field', 'string : 25359')])

    def test10_object_schema_to_line_tuples_long(self):
        res = object_schema_to_line_tuples(
            self.long_schema['object'],
            ['Field_full_name', 'Field_compact_name', 'Field_name', 'Type', 'Count', 'Types_count'],
            'foo.bar:')
        exp = [
            ('foo.bar:field', ' .  : field', 'field', 'string', 25359, 'string : 25359'),
            ('foo.bar:field2', ' .  : field2', 'field2', 'string', 25359, 'string : 25359'),
            ('foo.bar:field3', ' .  : field3', 'field3', 'ARRAY(OBJECT)', 25359,
             'ARRAY(OBJECT : 25359) : 25359'),
            ('foo.bar:field3:subfield1', ' .  :  : subfield1', 'subfield1', 'string', 25359,
             'string : 25359'),
            ('foo.bar:field3:subfield2', ' .  :  : subfield2', 'subfield2', 'ARRAY(string)',
             25359, 'object : 25359')]
        self.assertEqual(res, exp)

    def test11_mongo_schema_as_dataframe_simple(self):
        columns = ['Field_compact_name', 'Types_count']
        res = mongo_schema_as_dataframe(self.simple_full_schema, columns)
        exp = pd.DataFrame([['db', 'coll', 'field', 'string : 25359']],
                           columns=['Database', 'Collection'] + columns)
        assert_frame_equal(res, exp)

    def test12_mongo_schema_as_dataframe_long(self):
        columns = ['Field_full_name', 'Field_compact_name', 'Field_name', 'Type', 'Count',
                   'Types_count']
        res = mongo_schema_as_dataframe(self.long_full_schema, columns)
        exp = [['db1', 'coll', 'field', 'field', 'field', 'string', 25359, 'string : 25359'],
               ['db1', 'coll', 'field2', 'field2', 'field2', 'string', 25359, 'string : 25359'],
               ['db1', 'coll', 'field3', 'field3', 'field3', 'ARRAY(OBJECT)', 25359,
                'ARRAY(OBJECT : 25359) : 25359'],
               ['db1', 'coll', 'field3:subfield1', ' : subfield1', 'subfield1', 'string', 25359,
                'string : 25359'],
               ['db1', 'coll', 'field3:subfield2', ' : subfield2', 'subfield2', 'ARRAY(string)',
                25359, 'object : 25359'],
               ['db2', 'coll1', 'field', 'field', 'field', 'string', 25359, 'string : 25359'],
               ['db2', 'coll2', 'field', 'field', 'field', 'string', 25359, 'string : 25359']]
        exp = pd.DataFrame(exp, columns=['Database', 'Collection'] + columns)
        assert_frame_equal(res, exp)


class TestExportIntegration(TestRemovingOutputOnSuccess):
    @classmethod
    def setUpClass(cls):
        super(TestExportIntegration, cls).setUpClass()
        with open(os.path.join(TEST_DIR, 'resources', 'input', 'mapping.json')) as f:
            cls.mapping_ex_dict = json.loads(f.read())
        with open(os.path.join(TEST_DIR, 'resources', 'input', 'test_schema.json')) as f:
            cls.schema_ex_dict = json.loads(f.read())
        cls.columns = ['Field_compact_name', 'Field_name', 'Full_name', 'Description', 'Count',
                       'Percentage', 'Types_count']
        cls.schema_ex_df = mongo_schema_as_dataframe(cls.schema_ex_dict, cls.columns)

    def test01_write_txt(self):
        self.output = os.path.join(TEST_DIR, 'output_data_dict.txt')
        expected_file = os.path.join(TEST_DIR, 'resources', 'expected', 'data_dict.txt')
        with open(self.output, 'w') as out_fd:
            write_mongo_df_as_txt(self.schema_ex_df, out_fd)
        self.assertTrue(filecmp.cmp(self.output, expected_file))

    def test02_write_html(self):
        self.output = os.path.join(TEST_DIR, 'output_data_dict.html')
        expected_file = os.path.join(TEST_DIR, 'resources', 'expected', 'data_dict.html')
        with open(self.output, 'w') as out_fd:
            write_mongo_df_as_html(self.schema_ex_df, out_fd)
        with open(self.output) as out_fd, open(expected_file) as exp_fd:
            self.assertEqual(out_fd.read().replace(' ', ''), exp_fd.read().replace(' ', ''))

    def test03_write_xlsx(self):
        self.output = os.path.join(TEST_DIR, 'output_data_dict.xlsx')
        expected_file = os.path.join(TEST_DIR, 'resources', 'expected', 'data_dict.xlsx')
        write_mongo_df_as_xlsx(self.schema_ex_df, self.output)
        res = [cell.value for row in load_workbook(self.output).active for cell in row]
        exp = [cell.value for row in load_workbook(expected_file).active for cell in row]
        self.assertEqual(res, exp)

    def test04_write_output_dict_schema_txt(self):
        self.output = os.path.join(TEST_DIR, 'output_data_dict_from_schema.txt')
        expected_file = os.path.join(TEST_DIR, 'resources', 'expected', 'data_dict.txt')
        arg = {'--format': ['txt', 'txt'], '--output': self.output,
               '--columns': " ".join(self.columns)}
        write_output_dict(self.schema_ex_dict, arg)
        self.assertTrue(filecmp.cmp(self.output, expected_file))

    def test05_write_output_dict_schema_html(self):
        self.output = os.path.join(TEST_DIR, 'output_data_dict_from_schema.html')
        expected_file = os.path.join(TEST_DIR, 'resources', 'expected', 'data_dict.html')
        arg = {'--format': ['html'], '--output': self.output,
               '--columns': " ".join(self.columns)}
        write_output_dict(self.schema_ex_dict, arg)
        with open(self.output) as out_fd, open(expected_file) as exp_fd:
            self.assertEqual(out_fd.read().replace(' ', ''), exp_fd.read().replace(' ', ''))

    def test05_write_output_dict_schema_xlsx(self):
        self.output = os.path.join(TEST_DIR, 'output_data_dict_from_schema.xlsx')
        expected_file = os.path.join(TEST_DIR, 'resources', 'expected', 'data_dict.xlsx')
        arg = {'--format': ['xlsx'], '--output': self.output,
               '--columns': " ".join(self.columns)}
        write_output_dict(self.schema_ex_dict, arg)
        res = [cell.value for row in load_workbook(self.output).active for cell in row]
        exp = [cell.value for row in load_workbook(expected_file).active for cell in row]
        self.assertEqual(res, exp)

    def test06_write_output_dict_mapping_yaml(self):
        self.output = os.path.join(TEST_DIR, 'output_data_dict_from_mapping.yaml')
        expected_file = os.path.join(TEST_DIR, 'resources', 'expected', 'mapping.yaml')
        arg = {'--format': ['yaml'], '--output': self.output,
               '--columns': None, '--without-counts': True}
        write_output_dict(self.mapping_ex_dict, arg)
        self.assertTrue(filecmp.cmp(self.output, expected_file))

    def test07_write_output_dict_wrong_format(self):
        arg = {'--format': ['fake'], '--output': self.output,
               '--columns': None, '--without-counts': True}
        with self.assertRaises(ValueError):
            write_output_dict(self.mapping_ex_dict, arg)


if __name__ == '__main__':
    unittest.main()
