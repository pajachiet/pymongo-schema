import filecmp
import json
import os
import pytest
from itertools import chain

from openpyxl import load_workbook
from pymongo import MongoClient

from pymongo_schema.extract import extract_pymongo_client_schema
from pymongo_schema.tosql import mongo_schema_to_mapping
from pymongo_schema.__main__ import main

TEST_DIR = os.path.dirname(__file__)
SCHEMA_FILE = os.path.join(TEST_DIR, 'resources', 'input', 'test_schema.json')


@pytest.yield_fixture(scope='module')
def pymongo_client():
    conn = MongoClient()
    try:
        yield conn
    finally:
        conn.close()
        

def test00_from_mongo_to_mapping(pymongo_client):
    with open(os.path.join(TEST_DIR, "resources", "input", "mapping.json")) as f:
        exp_mapping = json.load(f)
    mongo_schema = extract_pymongo_client_schema(pymongo_client,
                                                 database_names='test_db',
                                                 collection_names='test_col')

    mapping = mongo_schema_to_mapping(mongo_schema)
    assert mapping == exp_mapping


def test00_from_mongo_to_mapping_long(pymongo_client):
    with open(os.path.join(TEST_DIR, 'resources', "functional", "exp_mapping.json")) as f:
        exp_mapping = json.load(f)
    mongo_schema = extract_pymongo_client_schema(pymongo_client,
                                                 database_names=['test_db', 'test_db1', 'test_db2'])

    mapping = mongo_schema_to_mapping(mongo_schema)
    assert mapping == exp_mapping


def test01_extract():
    output = os.path.join(TEST_DIR, "output_fctl_schema.json")
    expected_file = os.path.join(TEST_DIR, 'resources', 'expected', 'schema.json')
    argv = ['extract', '--database', 'test_db', '--collection', 'test_col',
            '--output', output, '--format', 'json']
    main(argv)
    with open(output) as out_f, open(expected_file) as exp_f:
        assert json.load(out_f) == json.load(exp_f)
    os.remove(output)


def test02_transform():
    base_output = "output_fctl_data_dict"
    outputs = {}
    extensions = ['html', 'xlsx', 'tsv', 'md']
    for ext in extensions:
        outputs[ext] = "{}.{}".format(base_output, ext)

    exp = os.path.join(TEST_DIR, 'resources', 'expected', 'data_dict')
    argv = ['transform', '--input', SCHEMA_FILE, '--output', base_output, '--columns',
            'Field_compact_name Field_name Full_name Description Count Percentage Types_count']
    argv += chain.from_iterable([['--format', fmt] for fmt in extensions])
    main(argv)

    assert filecmp.cmp(outputs['tsv'], "{}.tsv".format(exp))
    assert filecmp.cmp(outputs['md'], "{}.md".format(exp))
    with open(outputs['html']) as out_fd, \
            open("{}.html".format(exp)) as exp_fd:
        assert out_fd.read().replace(' ', '') == exp_fd.read().replace(' ', '')
    res = [cell.value for row in load_workbook(outputs['xlsx']).active for cell in row]
    exp = [cell.value for row in load_workbook("{}.xlsx".format(exp)).active for cell in row]
    assert res == exp
    for output in outputs.values():
        os.remove(output)


def test03_transform_filter():
    base_output = "output_fctl_data_dict_filtered"
    outputs = {}
    extensions = ['html', 'xlsx', 'tsv', 'md']
    for ext in extensions:
        outputs[ext] = "{}.{}".format(base_output, ext)

    namespace = os.path.join(TEST_DIR, 'resources', 'input', 'namespace.json')
    exp = os.path.join(TEST_DIR, 'resources', 'expected', 'data_dict_filtered')
    argv = ['transform', '--input', SCHEMA_FILE, '--output', base_output,
            '--filter', namespace, '--columns',
            'Field_compact_name Field_name Full_name Description Count Percentage Types_count']
    argv += chain.from_iterable([['--format', fmt] for fmt in extensions])
    main(argv)

    assert filecmp.cmp(outputs['tsv'], "{}.tsv".format(exp))
    assert filecmp.cmp(outputs['md'], "{}.md".format(exp))
    with open(outputs['html']) as out_fd, \
            open("{}.html".format(exp)) as exp_fd:
        assert out_fd.read().replace(' ', '') == exp_fd.read().replace(' ', '')
    res = [cell.value for row in load_workbook(outputs['xlsx']).active for cell in row]
    exp = [cell.value for row in load_workbook("{}.xlsx".format(exp)).active for cell in row]
    assert res == exp
    for output in outputs.values():
        os.remove(output)


def test04_transform_default_cols():
    base_output = "output_fctl_data_dict_default"
    outputs = {}
    extensions = ['html', 'xlsx', 'tsv', 'md']
    for ext in extensions:
        outputs[ext] = "{}.{}".format(base_output, ext)

    exp = os.path.join(TEST_DIR, 'resources', 'expected', 'data_dict_default')
    argv = ['transform', '--input', SCHEMA_FILE, '--output', base_output]
    argv += chain.from_iterable([['--format', fmt] for fmt in extensions])
    main(argv)

    assert filecmp.cmp(outputs['tsv'], "{}.tsv".format(exp))
    assert filecmp.cmp(outputs['md'], "{}.md".format(exp))
    with open(outputs['html']) as out_fd, \
            open("{}.html".format(exp)) as exp_fd:
        assert out_fd.read().replace(' ', '') == exp_fd.read().replace(' ', '')
    res = [cell.value for row in load_workbook(outputs['xlsx']).active for cell in row]
    exp = [cell.value for row in load_workbook("{}.xlsx".format(exp)).active for cell in row]
    assert res == exp
    for output in outputs.values():
        os.remove(output)


def test05_tosql():
    output = os.path.join(TEST_DIR, "output_fctl_mapping.json")
    exp = os.path.join(TEST_DIR, 'resources', 'expected', 'mapping.json')

    argv = ['tosql', '--input', SCHEMA_FILE, '--output', output]
    main(argv)

    with open(output) as out_fd, open(exp) as exp_fd:
        assert json.load(out_fd) == json.load(exp_fd)
    os.remove(output)


def test06_compare():
    base_output = "output_fctl_diff"
    outputs = {}
    extensions = ['html', 'xlsx', 'tsv', 'md']
    for ext in extensions:
        outputs[ext] = "{}.{}".format(base_output, ext)

    exp = os.path.join(TEST_DIR, 'resources', 'functional', 'expected', 'diff')
    exp_schema = os.path.join(TEST_DIR, 'resources', 'input', 'test_schema2.json')
    argv = ['compare', '--input', SCHEMA_FILE, '--output', base_output, '--expected', exp_schema]
    argv += chain.from_iterable([['--format', fmt] for fmt in extensions])
    main(argv)

    assert filecmp.cmp(outputs['tsv'], "{}.tsv".format(exp))
    assert filecmp.cmp(outputs['md'], "{}.md".format(exp))
    with open(outputs['html']) as out_fd, \
            open("{}.html".format(exp)) as exp_fd:
        assert out_fd.read().replace(' ', '') == exp_fd.read().replace(' ', '')
    res = [cell.value for row in load_workbook(outputs['xlsx']).active for cell in row]
    exp = [cell.value for row in load_workbook("{}.xlsx".format(exp)).active for cell in row]
    assert res == exp
    for output in outputs.values():
        os.remove(output)